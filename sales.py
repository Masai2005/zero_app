"""
Sales window module for the application.
"""
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
    QTableWidget, QTableWidgetItem, QComboBox, QDoubleSpinBox, 
    QSpinBox, QLineEdit, QFormLayout, QGroupBox, QMessageBox,
    QHeaderView, QDialog, QDateEdit, QTextEdit, QFileDialog,
    QProgressBar, QTabWidget
)
from PySide6.QtCore import Qt, Signal, QDate
from datetime import datetime, timedelta
import csv
import json
from collections import defaultdict, Counter

# PDF Export
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Excel Export
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from styles import StyleSheet, Theme, Colors
from utils import DataManager, format_currency, format_date, DEFAULT_SALES_FILE, DEFAULT_PRODUCTS_FILE, DEFAULT_CUSTOMERS_FILE, show_message, show_validation_error, handle_data_error, USER_TYPE_ADMIN, MovementManager
from validation import CustomerValidator, SaleValidator
from barcode_utils import BarcodeScannerDialog
from search_filter import SearchFilter, AdvancedSearchDialog, QuickSearchWidget
from print_utils import PrintManager


class NewCustomerDialog(QDialog):
    """Dialog for adding a new customer"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add New Customer")
        self.setMinimumWidth(400)
        self.setup_ui()
    
    def setup_ui(self):
        """Setup UI components"""
        layout = QVBoxLayout()
        
        # Form layout for inputs
        form_layout = QFormLayout()
        
        # First name field
        self.first_name_input = QLineEdit()
        form_layout.addRow("First Name:", self.first_name_input)
        
        # Middle name field
        self.middle_name_input = QLineEdit()
        form_layout.addRow("Middle Name:", self.middle_name_input)
        
        # Last name field
        self.last_name_input = QLineEdit()
        form_layout.addRow("Last Name:", self.last_name_input)
        
        # Company name field
        self.company_name_input = QLineEdit()
        form_layout.addRow("Company Name:", self.company_name_input)
        
        # Mobile number field
        self.mobile_input = QLineEdit()
        form_layout.addRow("Mobile:", self.mobile_input)
        
        # Email field
        self.email_input = QLineEdit()
        form_layout.addRow("Email:", self.email_input)
        
        # WhatsApp field
        self.whatsapp_input = QLineEdit()
        form_layout.addRow("WhatsApp:", self.whatsapp_input)
        
        # Address field
        self.address_input = QLineEdit()
        form_layout.addRow("Address:", self.address_input)
        
        layout.addLayout(form_layout)
        
        # Button layout
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        # Add stretch to push buttons to the right
        button_layout.addStretch()
        
        # Cancel button
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.setMinimumWidth(100)
        self.cancel_button.clicked.connect(self.reject)
        
        # Save button
        self.save_button = QPushButton("Save")
        self.save_button.setMinimumWidth(100)
        self.save_button.clicked.connect(self.save_customer)
        
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.save_button)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def save_customer(self):
        """Save new customer data with validation"""
        # Create customer object
        customer_data = {
            "id": DataManager.generate_id(),
            "first_name": self.first_name_input.text().strip(),
            "middle_name": self.middle_name_input.text().strip(),
            "last_name": self.last_name_input.text().strip(),
            "company_name": self.company_name_input.text().strip(),
            "mobile": self.mobile_input.text().strip(),
            "email": self.email_input.text().strip(),
            "whatsapp": self.whatsapp_input.text().strip(),
            "address": self.address_input.text().strip(),
            "created_at": datetime.now().isoformat()
        }
        
        # Validate customer data
        is_valid, error_message = CustomerValidator.validate_customer_data(customer_data)
        if not is_valid:
            show_message(self, "Validation Error", error_message, QMessageBox.Warning)
            return
        
        self.customer_data = customer_data
        self.accept()


class SalesHistoryDialog(QDialog):
    """Dialog for viewing sales history"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Sales History")
        self.setMinimumSize(800, 500)
        
        # Load sales data
        self.sales_data, error = DataManager.load_data(DEFAULT_SALES_FILE)
        if error:
            handle_data_error(self, "load sales history", error)
            self.sales_data = []
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup UI components"""
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("Sales History")
        title.setStyleSheet("font-size: 18px; font-weight: bold;")
        layout.addWidget(title)
        
        # Search section
        search_layout = QHBoxLayout()
        
        # Quick search widget
        self.quick_search = QuickSearchWidget("Search sales...")
        self.quick_search.search_changed.connect(self.on_quick_search)
        self.quick_search.advanced_search_requested.connect(self.show_advanced_search)
        search_layout.addWidget(self.quick_search)
        
        layout.addLayout(search_layout)
        
        # Sales table
        self.sales_table = QTableWidget()
        self.sales_table.setColumnCount(6)
        self.sales_table.setHorizontalHeaderLabels(["Date", "Invoice #", "Customer", "Items", "Total", "Payment Method"])
        self.sales_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.sales_table.selectionBehavior = QTableWidget.SelectRows
        self.sales_table.itemSelectionChanged.connect(self.on_selection_changed)
        layout.addWidget(self.sales_table)
        
        # Initialize search filters
        self.current_filters = {}
        self.filtered_sales = self.sales_data.copy()
        
        # Populate table
        self.populate_table()
        
        # Button layout
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        # Print receipt button
        self.print_receipt_btn = QPushButton("Print Receipt")
        self.print_receipt_btn.setMinimumWidth(120)
        self.print_receipt_btn.clicked.connect(self.print_selected_receipt)
        self.print_receipt_btn.setEnabled(False)
        button_layout.addWidget(self.print_receipt_btn)
        
        # Print invoice button
        self.print_invoice_btn = QPushButton("Print Invoice")
        self.print_invoice_btn.setMinimumWidth(120)
        self.print_invoice_btn.clicked.connect(self.print_selected_invoice)
        self.print_invoice_btn.setEnabled(False)
        button_layout.addWidget(self.print_invoice_btn)
        
        button_layout.addStretch()
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.setMinimumWidth(80)
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def on_quick_search(self, search_text):
        """Handle quick search text change"""
        if search_text.strip():
            self.current_filters = {"text_search": search_text.strip()}
        else:
            self.current_filters = {}
        self.apply_filters()
    
    def show_advanced_search(self):
        """Show advanced search dialog"""
        dialog = AdvancedSearchDialog("sales", self)
        
        # Pre-populate with current filters
        if self.current_filters:
            dialog.apply_config(self.current_filters)
        
        dialog.search_applied.connect(self.on_advanced_search_applied)
        dialog.exec()
    
    def on_advanced_search_applied(self, filters):
        """Handle advanced search filters"""
        self.current_filters = filters
        self.apply_filters()
        
        # Update quick search text if it was part of the filters
        if filters.get("text_search"):
            self.quick_search.set_search_text(filters["text_search"])
        else:
            self.quick_search.clear_search()
    
    def apply_filters(self):
        """Apply current filters to sales"""
        self.filtered_sales = SearchFilter.filter_sales(self.sales_data, self.current_filters)
        self.populate_table()
    
    def populate_table(self):
        """Populate the sales table with filtered data"""
        # Use filtered sales data
        sales_to_show = self.filtered_sales if hasattr(self, 'filtered_sales') else self.sales_data
        
        self.sales_table.setRowCount(len(sales_to_show))
        
        for row, sale in enumerate(sales_to_show):
            # Date
            self.sales_table.setItem(row, 0, QTableWidgetItem(format_date(sale.get("created_at", ""))))
            
            # Invoice number
            self.sales_table.setItem(row, 1, QTableWidgetItem(sale.get("invoice_number", "")))
            
            # Customer
            customer_name = "Walk-in Customer"
            if sale.get("customer_id"):
                customer_name = sale.get("customer_name", "Unknown")
            self.sales_table.setItem(row, 2, QTableWidgetItem(customer_name))
            
            # Items count
            items_count = len(sale.get("items", []))
            self.sales_table.setItem(row, 3, QTableWidgetItem(str(items_count)))
            
            # Total
            self.sales_table.setItem(row, 4, QTableWidgetItem(format_currency(sale.get("total", 0))))
            
            # Payment method
            self.sales_table.setItem(row, 5, QTableWidgetItem(sale.get("payment_method", "")))
    
    def on_selection_changed(self):
        """Handle table selection change"""
        has_selection = len(self.sales_table.selectedItems()) > 0
        self.print_receipt_btn.setEnabled(has_selection)
        self.print_invoice_btn.setEnabled(has_selection)
    
    def get_selected_sale(self):
        """Get the currently selected sale data"""
        current_row = self.sales_table.currentRow()
        if current_row >= 0 and current_row < len(self.filtered_sales):
            return self.filtered_sales[current_row]
        return None
    
    def print_selected_receipt(self):
        """Print receipt for selected sale"""
        sale_data = self.get_selected_sale()
        if sale_data:
            PrintManager.print_receipt(sale_data, self)
        else:
            show_message(self, "No Selection", "Please select a sale to print receipt.", QMessageBox.Warning)
    
    def print_selected_invoice(self):
        """Print invoice for selected sale"""
        sale_data = self.get_selected_sale()
        if sale_data:
            PrintManager.print_invoice(sale_data, self)
        else:
            show_message(self, "No Selection", "Please select a sale to print invoice.", QMessageBox.Warning)


class SalesReportDialog(QDialog):
    """Dialog for generating comprehensive sales reports"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Sales Report Generator")
        self.setMinimumSize(900, 700)
        
        # Load data
        self.sales_data, sales_error = DataManager.load_data(DEFAULT_SALES_FILE)
        if sales_error:
            handle_data_error(self, "load sales data", sales_error)
            self.sales_data = []
        
        self.customers_data, customers_error = DataManager.load_data(DEFAULT_CUSTOMERS_FILE)
        if customers_error:
            handle_data_error(self, "load customers data", customers_error)
            self.customers_data = []
        
        self.products_data, products_error = DataManager.load_data(DEFAULT_PRODUCTS_FILE)
        if products_error:
            handle_data_error(self, "load products data", products_error)
            self.products_data = []
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup UI components"""
        layout = QVBoxLayout()
        
        # Title
        title = QLabel("Sales Report Generator")
        title.setStyleSheet("font-size: 20px; font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(title)
        
        # Date range selection
        date_group = QGroupBox("Report Period")
        date_layout = QHBoxLayout()
        
        # From date
        date_layout.addWidget(QLabel("From:"))
        self.from_date = QDateEdit()
        self.from_date.setDate(QDate.currentDate().addDays(-30))  # Default to last 30 days
        self.from_date.setCalendarPopup(True)
        date_layout.addWidget(self.from_date)
        
        # To date
        date_layout.addWidget(QLabel("To:"))
        self.to_date = QDateEdit()
        self.to_date.setDate(QDate.currentDate())
        self.to_date.setCalendarPopup(True)
        date_layout.addWidget(self.to_date)
        
        # Generate button
        self.generate_btn = QPushButton("Generate Report")
        self.generate_btn.clicked.connect(self.generate_report)
        date_layout.addWidget(self.generate_btn)
        
        date_layout.addStretch()
        date_group.setLayout(date_layout)
        layout.addWidget(date_group)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Report tabs
        self.report_tabs = QTabWidget()
        layout.addWidget(self.report_tabs)
        
        # Export buttons
        export_layout = QHBoxLayout()
        export_layout.setSpacing(10)
        
        self.export_pdf_btn = QPushButton("Export to PDF")
        self.export_pdf_btn.setMinimumWidth(120)
        self.export_pdf_btn.clicked.connect(self.export_to_pdf)
        self.export_pdf_btn.setEnabled(False)
        export_layout.addWidget(self.export_pdf_btn)
        
        self.export_xlsx_btn = QPushButton("Export to Excel")
        self.export_xlsx_btn.setMinimumWidth(120)
        self.export_xlsx_btn.clicked.connect(self.export_to_xlsx)
        self.export_xlsx_btn.setEnabled(False)
        export_layout.addWidget(self.export_xlsx_btn)
        
        # Print report button
        self.print_report_btn = QPushButton("Print Report")
        self.print_report_btn.setMinimumWidth(120)
        self.print_report_btn.clicked.connect(self.print_report)
        self.print_report_btn.setEnabled(False)
        export_layout.addWidget(self.print_report_btn)
        
        export_layout.addStretch()
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.setMinimumWidth(80)
        close_btn.clicked.connect(self.accept)
        export_layout.addWidget(close_btn)
        
        layout.addLayout(export_layout)
        
        self.setLayout(layout)
        
        # Generate initial report
        self.generate_report()
    
    def format_currency(self, value):
        """Format currency values for display"""
        try:
            return f"{float(value):,.2f}"
        except (ValueError, TypeError):
            return "0.00"

    def filter_sales_by_date(self, from_date, to_date):
        """Filter sales data by date range"""
        filtered_sales = []
        
        for sale in self.sales_data:
            try:
                sale_date = datetime.fromisoformat(sale.get("created_at", ""))
                if from_date <= sale_date <= to_date:
                    filtered_sales.append(sale)
            except (ValueError, TypeError):
                continue
        
        return filtered_sales
    
    def generate_report(self):
        """Generate comprehensive sales report"""
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        # Get date range
        from_date = self.from_date.date().toPython()
        to_date = self.to_date.date().toPython()
        
        # Convert to datetime for comparison
        from_datetime = datetime.combine(from_date, datetime.min.time())
        to_datetime = datetime.combine(to_date, datetime.max.time())
        
        # Filter sales data
        filtered_sales = self.filter_sales_by_date(from_datetime, to_datetime)
        
        self.progress_bar.setValue(20)
        
        # Clear existing tabs
        self.report_tabs.clear()
        
        # Generate different report sections
        self.generate_summary_tab(filtered_sales)
        self.progress_bar.setValue(40)
        
        self.generate_top_products_tab(filtered_sales)
        self.progress_bar.setValue(60)
        
        self.generate_customer_analysis_tab(filtered_sales)
        self.progress_bar.setValue(80)
        
        self.generate_payment_methods_tab(filtered_sales)
        self.progress_bar.setValue(100)
        
        # Enable export buttons
        self.export_pdf_btn.setEnabled(True)
        self.export_xlsx_btn.setEnabled(True)
        
        # Store filtered data for export
        self.current_report_data = filtered_sales
        
        self.progress_bar.setVisible(False)
    
    def generate_summary_tab(self, sales_data):
        """Generate summary statistics tab"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Calculate summary statistics
        total_sales = len(sales_data)
        total_revenue = sum(sale.get("total", 0) for sale in sales_data)
        total_discount = sum(sale.get("discount", 0) for sale in sales_data)
        average_sale = total_revenue / total_sales if total_sales > 0 else 0
        
        # Count unique customers
        unique_customers = set()
        for sale in sales_data:
            if sale.get("customer_id"):
                unique_customers.add(sale.get("customer_id"))
        
        # Summary text
        summary_text = f"""
SALES SUMMARY REPORT
Period: {self.from_date.date().toString()} to {self.to_date.date().toString()}

KEY METRICS:
• Total Sales Transactions: {total_sales:,}
• Total Revenue: {format_currency(total_revenue)}
• Total Discounts Given: {format_currency(total_discount)}
• Average Sale Value: {format_currency(average_sale)}
• Unique Customers Served: {len(unique_customers)}

DAILY AVERAGES:
• Average Transactions per Day: {total_sales / max(1, (self.to_date.date().toPython() - self.from_date.date().toPython()).days + 1):.1f}
• Average Revenue per Day: {format_currency(total_revenue / max(1, (self.to_date.date().toPython() - self.from_date.date().toPython()).days + 1))}
        """
        
        summary_display = QTextEdit()
        summary_display.setPlainText(summary_text.strip())
        summary_display.setReadOnly(True)
        summary_display.setFont(summary_display.font())
        
        layout.addWidget(summary_display)
        tab.setLayout(layout)
        
        self.report_tabs.addTab(tab, "Summary")
    
    def generate_top_products_tab(self, sales_data):
        """Generate top products analysis tab"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Analyze product sales
        product_stats = defaultdict(lambda: {"quantity": 0, "revenue": 0, "transactions": 0})
        
        for sale in sales_data:
            for item in sale.get("items", []):
                product_id = item.get("product_id")
                product_name = item.get("product_name", "Unknown")
                quantity = item.get("quantity", 0)
                total_price = item.get("total_price", 0)
                
                key = f"{product_name} (ID: {product_id})"
                product_stats[key]["quantity"] += quantity
                product_stats[key]["revenue"] += total_price
                product_stats[key]["transactions"] += 1
        
        # Create table
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Product", "Quantity Sold", "Revenue", "Transactions"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Sort by revenue (descending)
        sorted_products = sorted(product_stats.items(), key=lambda x: x[1]["revenue"], reverse=True)
        
        table.setRowCount(len(sorted_products))
        
        for row, (product, stats) in enumerate(sorted_products):
            table.setItem(row, 0, QTableWidgetItem(product))
            table.setItem(row, 1, QTableWidgetItem(str(stats["quantity"])))
            table.setItem(row, 2, QTableWidgetItem(format_currency(stats["revenue"])))
            table.setItem(row, 3, QTableWidgetItem(str(stats["transactions"])))
        
        layout.addWidget(QLabel("Top Products by Revenue"))
        layout.addWidget(table)
        tab.setLayout(layout)
        
        self.report_tabs.addTab(tab, "Top Products")
    
    def generate_customer_analysis_tab(self, sales_data):
        """Generate customer analysis tab"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Analyze customer data
        customer_stats = defaultdict(lambda: {"transactions": 0, "revenue": 0, "items": 0})
        walk_in_stats = {"transactions": 0, "revenue": 0, "items": 0}
        
        for sale in sales_data:
            customer_id = sale.get("customer_id")
            customer_name = sale.get("customer_name", "Walk-in Customer")
            total = sale.get("total", 0)
            items_count = len(sale.get("items", []))
            
            if customer_id:
                customer_stats[customer_name]["transactions"] += 1
                customer_stats[customer_name]["revenue"] += total
                customer_stats[customer_name]["items"] += items_count
            else:
                walk_in_stats["transactions"] += 1
                walk_in_stats["revenue"] += total
                walk_in_stats["items"] += items_count
        
        # Create table
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Customer", "Transactions", "Total Revenue", "Items Purchased"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Sort by revenue (descending)
        sorted_customers = sorted(customer_stats.items(), key=lambda x: x[1]["revenue"], reverse=True)
        
        # Add walk-in customers at the end
        if walk_in_stats["transactions"] > 0:
            sorted_customers.append(("Walk-in Customers", walk_in_stats))
        
        table.setRowCount(len(sorted_customers))
        
        for row, (customer, stats) in enumerate(sorted_customers):
            table.setItem(row, 0, QTableWidgetItem(customer))
            table.setItem(row, 1, QTableWidgetItem(str(stats["transactions"])))
            table.setItem(row, 2, QTableWidgetItem(format_currency(stats["revenue"])))
            table.setItem(row, 3, QTableWidgetItem(str(stats["items"])))
        
        layout.addWidget(QLabel("Customer Analysis"))
        layout.addWidget(table)
        tab.setLayout(layout)
        
        self.report_tabs.addTab(tab, "Customer Analysis")
    
    def generate_payment_methods_tab(self, sales_data):
        """Generate payment methods analysis tab"""
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Analyze payment methods
        payment_stats = defaultdict(lambda: {"transactions": 0, "revenue": 0})
        
        for sale in sales_data:
            payment_method = sale.get("payment_method", "Unknown")
            total = sale.get("total", 0)
            
            payment_stats[payment_method]["transactions"] += 1
            payment_stats[payment_method]["revenue"] += total
        
        # Create table
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Payment Method", "Transactions", "Total Revenue"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        # Sort by revenue (descending)
        sorted_payments = sorted(payment_stats.items(), key=lambda x: x[1]["revenue"], reverse=True)
        
        table.setRowCount(len(sorted_payments))
        
        for row, (method, stats) in enumerate(sorted_payments):
            table.setItem(row, 0, QTableWidgetItem(method))
            table.setItem(row, 1, QTableWidgetItem(str(stats["transactions"])))
            table.setItem(row, 2, QTableWidgetItem(format_currency(stats["revenue"])))
        
        layout.addWidget(QLabel("Payment Methods Analysis"))
        layout.addWidget(table)
        tab.setLayout(layout)
        
        self.report_tabs.addTab(tab, "Payment Methods")
    
    def export_to_pdf(self):
        """Export report data to PDF"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Export Sales Report to PDF", 
                f"sales_report_{self.from_date.date().toString('yyyy-MM-dd')}_to_{self.to_date.date().toString('yyyy-MM-dd')}.pdf",
                "PDF Files (*.pdf)"
            )
            
            if not file_path:
                return
                
            # Create document
            doc = SimpleDocTemplate(
                file_path,
                pagesize=landscape(letter),
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=30
            )
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30
            )
            
            # Content elements
            elements = []
            
            # Title
            title = Paragraph(
                f"Sales Report ({self.from_date.date().toString('yyyy-MM-dd')} to {self.to_date.date().toString('yyyy-MM-dd')})",
                title_style
            )
            elements.append(title)
            
            # Table data
            table_data = [
                ["Date", "Invoice Number", "Customer", "Payment Method", 
                 "Subtotal", "Discount", "Total", "Items", "Created By"]
            ]
            
            for sale in self.current_report_data:
                table_data.append([
                    format_date(sale.get("created_at", "")),
                    sale.get("invoice_number", ""),
                    sale.get("customer_name", "Walk-in Customer"),
                    sale.get("payment_method", ""),
                    self.format_currency(sale.get("subtotal", 0)),
                    self.format_currency(sale.get("discount", 0)),
                    self.format_currency(sale.get("total", 0)),
                    str(len(sale.get("items", []))),
                    sale.get("created_by", "")
                ])
            
            # Create table
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (4, 0), (6, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(table)
            
            # Build document
            doc.build(elements)
            
            show_message(self, "Export Successful", f"Report exported to {file_path}")
            
        except Exception as e:
            show_message(self, "Export Error", f"Failed to export report: {str(e)}", QMessageBox.Critical)
    
    def export_to_xlsx(self):
        """Export report data to Excel"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Export Sales Report to Excel", 
                f"sales_report_{self.from_date.date().toString('yyyy-MM-dd')}_to_{self.to_date.date().toString('yyyy-MM-dd')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Report"
            
            # Title
            ws.merge_cells('A1:I1')
            title_cell = ws['A1']
            title_cell.value = f"Sales Report ({self.from_date.date().toString('yyyy-MM-dd')} to {self.to_date.date().toString('yyyy-MM-dd')})"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Headers
            headers = ["Date", "Invoice Number", "Customer", "Payment Method", 
                      "Subtotal", "Discount", "Total", "Items Count", "Created By"]
            
            # Style for headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Add headers
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row, sale in enumerate(self.current_report_data, start=4):
                ws.cell(row=row, column=1, value=format_date(sale.get("created_at", "")))
                ws.cell(row=row, column=2, value=sale.get("invoice_number", ""))
                ws.cell(row=row, column=3, value=sale.get("customer_name", "Walk-in Customer"))
                ws.cell(row=row, column=4, value=sale.get("payment_method", ""))
                ws.cell(row=row, column=5, value=float(sale.get("subtotal", 0)))
                ws.cell(row=row, column=6, value=float(sale.get("discount", 0)))
                ws.cell(row=row, column=7, value=float(sale.get("total", 0)))
                ws.cell(row=row, column=8, value=len(sale.get("items", [])))
                ws.cell(row=row, column=9, value=sale.get("created_by", ""))
            
            # Format currency columns
            for col in [5, 6, 7]:  # Subtotal, Discount, Total columns
                for row in range(4, len(self.current_report_data) + 4):
                    cell = ws.cell(row=row, column=col)
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            # Adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            # Add spacing
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 15
            
            wb.save(file_path)
            
            show_message(self, "Export Successful", f"Report exported to {file_path}")
            
        except Exception as e:
            show_message(self, "Export Error", f"Failed to export report: {str(e)}", QMessageBox.Critical)
    
    def print_report(self):
        """Print the sales report"""
        try:
            if not hasattr(self, 'current_report_data') or not self.current_report_data:
                show_message(self, "No Data", "Please generate a report first before printing.", QMessageBox.Warning)
                return
            
            # Get date range string
            from_date = self.from_date.date().toString("yyyy-MM-dd")
            to_date = self.to_date.date().toString("yyyy-MM-dd")
            date_range = f"{from_date} to {to_date}"
            
            # Print the report
            PrintManager.print_sales_report(self.current_report_data, date_range, self)
            
        except Exception as e:
            show_message(self, "Print Error", f"Failed to print report: {str(e)}", QMessageBox.Critical)


class SalesWindow(QWidget):
    """Sales window widget"""
    
    def __init__(self, user_data, theme=Theme.LIGHT):
        super().__init__()
        self.user_data = user_data
        self.theme = theme
        self.cart = []
        self.customer_info = None
        
        # Load data with error handling
        self.products, products_error = DataManager.load_data(DEFAULT_PRODUCTS_FILE)
        if products_error:
            handle_data_error(self, "load products", products_error)
        
        self.customers, customers_error = DataManager.load_data(DEFAULT_CUSTOMERS_FILE)
        if customers_error:
            handle_data_error(self, "load customers", customers_error)
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup UI components"""
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)
        
        # Left side - Sales form
        left_layout = QVBoxLayout()
        
        # Title
        title = QLabel("Sales")
        title.setStyleSheet("font-size: 24px; font-weight: bold;")
        left_layout.addWidget(title)
        
        # Sales form
        sales_form = QGroupBox("New Sale")
        form_layout = QFormLayout()
        
        # Product selection with barcode scanning
        product_layout = QHBoxLayout()
        self.product_combo = QComboBox()
        self.refresh_products()
        product_layout.addWidget(self.product_combo)
        
        # Barcode scan button
        self.scan_barcode_btn = QPushButton("📷 Scan")
        self.scan_barcode_btn.setMinimumWidth(80)
        self.scan_barcode_btn.setMaximumWidth(90)
        self.scan_barcode_btn.setMinimumHeight(30)
        self.scan_barcode_btn.clicked.connect(self.scan_barcode)
        self.scan_barcode_btn.setToolTip("Scan barcode to select product")
        product_layout.addWidget(self.scan_barcode_btn)
        
        product_widget = QWidget()
        product_widget.setLayout(product_layout)
        form_layout.addRow("Product:", product_widget)
        
        # Unit of sale
        self.unit_combo = QComboBox()
        self.unit_combo.addItems(["Each", "Box", "Kg", "Liter"])
        form_layout.addRow("Unit:", self.unit_combo)
        
        # Quantity
        self.quantity_spin = QSpinBox()
        self.quantity_spin.setMinimum(1)
        self.quantity_spin.setMaximum(1000)
        form_layout.addRow("Quantity:", self.quantity_spin)
        
        # Discount
        self.discount_spin = QDoubleSpinBox()
        self.discount_spin.setMinimum(0)
        self.discount_spin.setMaximum(100)
        self.discount_spin.setSuffix("%")
        form_layout.addRow("Discount:", self.discount_spin)
        
        # Payment method
        self.payment_combo = QComboBox()
        self.payment_combo.addItems(["Cash", "Credit Card", "Credit (Account)"])
        self.payment_combo.currentTextChanged.connect(self.on_payment_method_changed)
        form_layout.addRow("Payment Method:", self.payment_combo)
        
        # Add to cart button
        self.add_to_cart_btn = QPushButton("Add to Cart")
        self.add_to_cart_btn.setMinimumWidth(120)
        self.add_to_cart_btn.setMinimumHeight(35)
        self.add_to_cart_btn.clicked.connect(self.add_to_cart)
        form_layout.addRow("", self.add_to_cart_btn)
        
        sales_form.setLayout(form_layout)
        left_layout.addWidget(sales_form)
        
        # Customer information section
        self.customer_info_group = QGroupBox("Customer Information")
        self.customer_info_group.setVisible(False)
        
        customer_form = QFormLayout()
        
        # Customer selection
        self.customer_combo = QComboBox()
        self.refresh_customers()
        customer_form.addRow("Customer:", self.customer_combo)
        
        # New customer button
        self.new_customer_btn = QPushButton("New Customer")
        self.new_customer_btn.setMinimumWidth(120)
        self.new_customer_btn.setMinimumHeight(30)
        self.new_customer_btn.clicked.connect(self.show_new_customer_form)
        customer_form.addRow("", self.new_customer_btn)
        
        self.customer_info_group.setLayout(customer_form)
        left_layout.addWidget(self.customer_info_group)
        
        # Add stretch to push everything to the top
        left_layout.addStretch()
        
        # Right side - Cart and checkout
        right_layout = QVBoxLayout()
        
        # Cart title
        cart_title = QLabel("Cart")
        cart_title.setStyleSheet("font-size: 18px; font-weight: bold;")
        right_layout.addWidget(cart_title)
        
        # Cart table
        self.cart_table = QTableWidget()
        self.cart_table.setColumnCount(5)
        self.cart_table.setHorizontalHeaderLabels(["Product", "Unit", "Quantity", "Price", "Total"])
        self.cart_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        right_layout.addWidget(self.cart_table)
        
        # Cart controls
        cart_controls = QHBoxLayout()
        cart_controls.setSpacing(10)
        
        # Remove from cart button
        self.remove_btn = QPushButton("Remove Selected")
        self.remove_btn.setMinimumWidth(140)
        self.remove_btn.clicked.connect(self.remove_from_cart)
        cart_controls.addWidget(self.remove_btn)
        
        # Clear cart button
        self.clear_btn = QPushButton("Clear All")
        self.clear_btn.setMinimumWidth(100)
        self.clear_btn.clicked.connect(self.clear_cart)
        cart_controls.addWidget(self.clear_btn)
        
        cart_controls.addStretch()
        
        right_layout.addLayout(cart_controls)
        
        # Total section
        totals = QFormLayout()
        
        # Subtotal
        self.subtotal_label = QLabel("$0.00")
        totals.addRow("Subtotal:", self.subtotal_label)
        
        # Discount
        self.total_discount_label = QLabel("$0.00")
        totals.addRow("Total Discount:", self.total_discount_label)
        
        # Total
        self.total_label = QLabel("$0.00")
        self.total_label.setStyleSheet("font-weight: bold;")
        totals.addRow("Total:", self.total_label)
        
        right_layout.addLayout(totals)
        
        # Checkout button
        self.checkout_btn = QPushButton("Complete Sale")
        self.checkout_btn.setMinimumHeight(40)
        self.checkout_btn.setMinimumWidth(150)
        self.checkout_btn.clicked.connect(self.complete_sale)
        self.checkout_btn.setStyleSheet(f"background-color: {Colors.LIGHT_SUCCESS if self.theme == Theme.LIGHT else Colors.DARK_SUCCESS}; color: white; font-weight: bold;")
        right_layout.addWidget(self.checkout_btn)
        
        # Sales history button
        self.history_btn = QPushButton("Sales History")
        self.history_btn.setMinimumHeight(35)
        self.history_btn.setMinimumWidth(130)
        self.history_btn.clicked.connect(self.show_sales_history)
        right_layout.addWidget(self.history_btn)
        
        # Sales report button (admin only)
        if self.user_data.get("type") == USER_TYPE_ADMIN:
            self.report_btn = QPushButton("Generate Sales Report")
            self.report_btn.setMinimumHeight(35)
            self.report_btn.setMinimumWidth(180)
            self.report_btn.clicked.connect(self.show_sales_report)
            self.report_btn.setStyleSheet(f"background-color: {Colors.LIGHT_PRIMARY if self.theme == Theme.LIGHT else Colors.DARK_PRIMARY}; color: white; font-weight: bold;")
            right_layout.addWidget(self.report_btn)
        
        # Add layouts to main layout
        main_layout.addLayout(left_layout, 1)
        main_layout.addLayout(right_layout, 2)
        
        self.setLayout(main_layout)
    
    def refresh_products(self):
        """Refresh the products combo box"""
        self.product_combo.clear()
        
        if self.products:
            for product in self.products:
                self.product_combo.addItem(product.get("name", "Unknown"), product)
    
    def refresh_customers(self):
        """Refresh the customers combo box"""
        self.customer_combo.clear()
        
        if self.customers:
            for customer in self.customers:
                name = f"{customer.get('first_name', '')} {customer.get('last_name', '')}"
                self.customer_combo.addItem(name, customer)
    
    def scan_barcode(self):
        """Open barcode scanner dialog"""
        scanner_dialog = BarcodeScannerDialog(self)
        scanner_dialog.barcode_scanned.connect(self.on_barcode_scanned)
        scanner_dialog.exec()
    
    def on_barcode_scanned(self, barcode_data):
        """Handle scanned barcode data"""
        # Find product by barcode
        matching_product = None
        for product in self.products:
            if product.get("barcode") == barcode_data:
                matching_product = product
                break
        
        if matching_product:
            # Select the matching product in the combo box
            for i in range(self.product_combo.count()):
                if self.product_combo.itemData(i) == matching_product:
                    self.product_combo.setCurrentIndex(i)
                    show_message(self, "Product Found", f"Product '{matching_product.get('name')}' selected from barcode scan.")
                    break
        else:
            # Show message that product was not found
            reply = QMessageBox.question(
                self, 
                "Product Not Found", 
                f"No product found with barcode '{barcode_data}'.\n\nWould you like to create a new product with this barcode?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                self.create_product_from_barcode(barcode_data)
    
    def create_product_from_barcode(self, barcode_data):
        """Create a new product with the scanned barcode"""
        # This would typically open the inventory management to add a new product
        # For now, we'll show a message directing the user to inventory management
        show_message(
            self, 
            "Create New Product", 
            f"To create a new product with barcode '{barcode_data}', please:\n\n"
            "1. Go to Inventory Management\n"
            "2. Click 'Add New Product'\n"
            "3. Enter the barcode: {barcode_data}\n"
            "4. Fill in the product details\n\n"
            "The product will then be available for selection in sales.",
            QMessageBox.Information
        )
    
    def on_payment_method_changed(self, method):
        """Handle payment method change"""
        if method == "Credit (Account)":
            self.customer_info_group.setVisible(True)
        else:
            self.customer_info_group.setVisible(False)
    
    def add_to_cart(self):
        """Add selected product to cart"""
        if self.product_combo.count() == 0:
            show_message(self, "Error", "No products available", QMessageBox.Warning)
            return
        
        product_data = self.product_combo.currentData()
        if not product_data:
            return
        
        quantity = self.quantity_spin.value()
        unit = self.unit_combo.currentText()
        discount_percent = self.discount_spin.value()
        
        # Check if quantity is valid
        if quantity <= 0:
            show_message(self, "Error", "Quantity must be greater than 0", QMessageBox.Warning)
            return
        
        # Check if there's enough stock
        if product_data.get("quantity", 0) < quantity:
            show_message(self, "Error", "Not enough stock available", QMessageBox.Warning)
            return
        
        # Calculate price
        unit_price = product_data.get("selling_price", 0)
        total_price = unit_price * quantity
        discount_amount = (total_price * discount_percent) / 100
        final_price = total_price - discount_amount
        
        # Add to cart
        cart_item = {
            "product_id": product_data.get("id"),
            "product_name": product_data.get("name"),
            "unit": unit,
            "quantity": quantity,
            "unit_price": unit_price,
            "discount_percent": discount_percent,
            "discount_amount": discount_amount,
            "total_price": final_price
        }
        
        self.cart.append(cart_item)
        self.update_cart_table()
        self.update_totals()
        
        # Reset input fields
        self.quantity_spin.setValue(1)
        self.discount_spin.setValue(0)
    
    def update_cart_table(self):
        """Update the cart table with current cart items"""
        self.cart_table.setRowCount(len(self.cart))
        
        for row, item in enumerate(self.cart):
            self.cart_table.setItem(row, 0, QTableWidgetItem(item.get("product_name")))
            self.cart_table.setItem(row, 1, QTableWidgetItem(item.get("unit")))
            self.cart_table.setItem(row, 2, QTableWidgetItem(str(item.get("quantity"))))
            
            # Price with discount indicator
            price_text = format_currency(item.get("unit_price"))
            if item.get("discount_percent") > 0:
                price_text += f" (-{item.get('discount_percent')}%)"
            self.cart_table.setItem(row, 3, QTableWidgetItem(price_text))
            
            # Total price
            self.cart_table.setItem(row, 4, QTableWidgetItem(format_currency(item.get("total_price"))))
    
    def update_totals(self):
        """Update the totals section"""
        subtotal = sum(item.get("quantity") * item.get("unit_price") for item in self.cart)
        total_discount = sum(item.get("discount_amount") for item in self.cart)
        total = subtotal - total_discount
        
        self.subtotal_label.setText(format_currency(subtotal))
        self.total_discount_label.setText(format_currency(total_discount))
        self.total_label.setText(format_currency(total))
    
    def remove_from_cart(self):
        """Remove selected item from cart"""
        selected_row = self.cart_table.currentRow()
        if selected_row >= 0:
            self.cart.pop(selected_row)
            self.update_cart_table()
            self.update_totals()
    
    def clear_cart(self):
        """Clear all items from cart"""
        self.cart = []
        self.update_cart_table()
        self.update_totals()
    
    def show_new_customer_form(self):
        """Show form to add a new customer"""
        dialog = NewCustomerDialog(self)
        if dialog.exec() == QDialog.Accepted:
            # Add new customer to data
            new_customer = dialog.customer_data
            self.customers.append(new_customer)
            
            # Save customer data
            success, error = DataManager.save_data(self.customers, DEFAULT_CUSTOMERS_FILE)
            if success:
                # Refresh customer combo box and select the new customer
                self.refresh_customers()
                self.customer_combo.setCurrentIndex(self.customer_combo.count() - 1)
            else:
                handle_data_error(self, "save customer data", error)
                # Remove the customer from memory since save failed
                self.customers.pop()
    
    def complete_sale(self):
        """Complete the sale"""
        if not self.cart:
            show_message(self, "Error", "Cart is empty", QMessageBox.Warning)
            return
        
        payment_method = self.payment_combo.currentText()
        
        # Check if customer is required
        if payment_method == "Credit (Account)" and self.customer_combo.count() == 0:
            show_message(self, "Error", "Credit sale requires a customer", QMessageBox.Warning)
            return
        
        # Get customer info if selected
        customer_id = None
        customer_name = "Walk-in Customer"
        
        if payment_method == "Credit (Account)":
            customer_data = self.customer_combo.currentData()
            if customer_data:
                customer_id = customer_data.get("id")
                customer_name = f"{customer_data.get('first_name', '')} {customer_data.get('last_name', '')}"
        
        # Calculate totals
        subtotal = sum(item.get("quantity") * item.get("unit_price") for item in self.cart)
        total_discount = sum(item.get("discount_amount") for item in self.cart)
        total = subtotal - total_discount
        
        # Generate invoice number
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        invoice_number = f"INV-{timestamp}"
        
        # Create sale record
        sale = {
            "id": DataManager.generate_id(),
            "invoice_number": invoice_number,
            "customer_id": customer_id,
            "customer_name": customer_name,
            "items": self.cart,
            "subtotal": subtotal,
            "discount": total_discount,
            "total": total,
            "payment_method": payment_method,
            "created_by": self.user_data.get("username"),
            "created_at": datetime.now().isoformat()
        }
        
        # Validate sale data before saving
        is_valid, validation_error = SaleValidator.validate_sale_data(sale)
        if not is_valid:
            show_message(self, "Validation Error", validation_error, QMessageBox.Warning)
            return
        
        # Load existing sales data
        sales_data, load_error = DataManager.load_data(DEFAULT_SALES_FILE)
        if load_error:
            handle_data_error(self, "load sales data", load_error)
            return
        
        # Add new sale
        sales_data.append(sale)
        
        # Update product quantities first
        products_updated = self.update_product_quantities()
        if not products_updated:
            return
        
        # Save sales data
        success, save_error = DataManager.save_data(sales_data, DEFAULT_SALES_FILE)
        if not success:
            handle_data_error(self, "save sale data", save_error)
            return
        
        # Create movement records for inventory tracking
        movement_success, movement_error = MovementManager.create_sale_movement(sale, self.user_data)
        if not movement_success:
            DataManager.logger.warning(f"Failed to create movement records: {movement_error}")
            # Don't fail the sale, just log the warning
        
        # Show success message and offer print options
        success_msg = QMessageBox(self)
        success_msg.setWindowTitle("Sale Completed")
        success_msg.setText(f"Sale completed successfully.\nInvoice Number: {invoice_number}")
        success_msg.setIcon(QMessageBox.Information)
        
        # Add custom buttons for print options
        print_receipt_btn = success_msg.addButton("Print Receipt", QMessageBox.ActionRole)
        print_invoice_btn = success_msg.addButton("Print Invoice", QMessageBox.ActionRole) 
        close_btn = success_msg.addButton("Close", QMessageBox.AcceptRole)
        
        success_msg.exec()
        
        # Handle print options
        if success_msg.clickedButton() == print_receipt_btn:
            PrintManager.print_receipt(sale, self)
        elif success_msg.clickedButton() == print_invoice_btn:
            PrintManager.print_invoice(sale, self)
        
        self.clear_cart()
    
    def update_product_quantities(self):
        """Update product quantities based on cart items"""
        try:
            # Load current products
            products, error = DataManager.load_data(DEFAULT_PRODUCTS_FILE)
            if error:
                handle_data_error(self, "load products for quantity update", error)
                return False
            
            # Create product lookup
            product_lookup = {p.get("id"): p for p in products}
            
            # Check stock availability and update quantities
            for item in self.cart:
                product_id = item.get("product_id")
                quantity_sold = item.get("quantity", 0)
                
                product = product_lookup.get(product_id)
                if not product:
                    show_message(self, "Error", f"Product not found: {item.get('product_name', 'Unknown')}", QMessageBox.Critical)
                    return False
                
                current_quantity = product.get("quantity", 0)
                if current_quantity < quantity_sold:
                    show_message(self, "Error", f"Insufficient stock for {product.get('name', 'Unknown')}.\nAvailable: {current_quantity}, Required: {quantity_sold}", QMessageBox.Critical)
                    return False
                
                # Update quantity
                product["quantity"] = current_quantity - quantity_sold
                product["updated_at"] = datetime.now().isoformat()
            
            # Save updated products
            success, save_error = DataManager.save_data(products, DEFAULT_PRODUCTS_FILE)
            if not success:
                handle_data_error(self, "update product quantities", save_error)
                return False
            
            # Update local products data
            self.products = products
            return True
            
        except Exception as e:
            error_msg = f"Error updating product quantities: {e}"
            DataManager.logger.error(error_msg)
            show_message(self, "Error", error_msg, QMessageBox.Critical)
            return False
    
    def show_sales_history(self):
        """Show sales history dialog"""
        dialog = SalesHistoryDialog(self)
        dialog.exec()
    
    def show_sales_report(self):
        """Show sales report dialog (admin only)"""
        dialog = SalesReportDialog(self)
        dialog.exec()
    
    def update_theme(self, theme):
        """Update the theme"""
        self.theme = theme
        # Update any theme-specific styling
        self.checkout_btn.setStyleSheet(f"background-color: {Colors.LIGHT_SUCCESS if self.theme == Theme.LIGHT else Colors.DARK_SUCCESS}; color: white;")
        
        # Update report button styling if it exists (admin only)
        if hasattr(self, 'report_btn'):
            self.report_btn.setStyleSheet(f"background-color: {Colors.LIGHT_PRIMARY if self.theme == Theme.LIGHT else Colors.DARK_PRIMARY}; color: white;")